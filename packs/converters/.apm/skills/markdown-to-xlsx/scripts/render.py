#!/usr/bin/env python3
"""
render.py — fill a branded Excel template from a Markdown artifact.

This is the deterministic renderer for the `markdown-to-xlsx` skill. The agent
assembles the content model and invokes this script; the script writes Markdown
content into the **data ranges** (named ranges and Excel Tables) a designer
already defined in a `.xlsx` template, rather than building a workbook from
scratch — so the template's formatting, formulas, and any chart that reads those
ranges survive.

Verbs
    --check                 import-probe openpyxl; exit 0 (present) / 2 (absent)
    inspect <template>      print the named ranges + Excel Tables (the fill-points)
    render  <markdown>      fill a template and write a .xlsx
        --template <path>   workbook with named ranges/tables; omit for default
        --output <path>     output path (default: <markdown-basename>.xlsx in CWD)

Stdout markers (for agent parsing)
    FILLPOINTS: kind=<defined_name|table> name=<n> ref=<ref>   (inspect)
    GUIDANCE: <msg>         workbook has no fill-points; how to add them
    OUTPUT: <path>          path to the written .xlsx
    FILLED: <n>             count of fill-points a value was written into
    WARNING: <msg>          non-fatal issue, surface to the user

Data-ranges-only contract
    The script writes only into named-range and Excel-Table data cells. It never
    creates, manipulates, or resizes chart/shape objects, and never resizes a
    table's range — so charts that read those ranges keep working. openpyxl
    preserves the charts and images it can parse on round-trip, but its tutorial
    warns that *shapes it cannot read are lost when an existing file is opened
    and saved*; for a template with complex Excel-authored drawings, re-open the
    result and confirm the visuals survived.

Trust model
    A user-supplied template is trusted-author input, consistent with the
    converters pack's local-files-trusted stance. openpyxl does not evaluate
    workbook content as code, so there is no template-injection surface here;
    XXE / zip-bomb on a crafted archive is an accepted, out-of-scope risk.
    Output-path confinement is still enforced (the output path is assembled from
    model-influenced content).

Errors go to stderr; exit code 1 on failure, 2 on a missing library.
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path
from typing import NamedTuple

PIP_INSTALL = "python -m pip install 'openpyxl>=3.1.0'"

GUIDANCE_NO_FILLPOINTS = (
    "this workbook has no named ranges or Excel Tables, so there are no "
    "fill-points to write into. In Excel, select the cell(s) for a scalar and "
    "give the selection a Name (Formulas → Define Name), and turn a data block "
    "into a Table (Insert → Table) for tabular content. Then re-run. The skill "
    "never silently converts an untagged workbook — that would discard your "
    "formatting and any chart wiring."
)


class RenderResult(NamedTuple):
    written: bool
    filled: int
    warnings: list[str]
    guidance: str | None


# --------------------------------------------------------------------------- #
# Output-path confinement
# --------------------------------------------------------------------------- #
def confine(path: Path, root: Path) -> Path:
    """Resolve ``path`` (following symlinks) and require it under ``root``.

    Path-*component* containment (``root`` is the resolved path or among its
    ``.parents``), not a string-prefix check, so a sibling like ``workdir-evil``
    is rejected against root ``workdir``. Raises ``ValueError`` on escape.
    """
    resolved = path.resolve()
    root_resolved = root.resolve()
    if resolved == root_resolved or root_resolved in resolved.parents:
        return resolved
    raise ValueError(
        f"path {path!r} resolves to {resolved} which is outside the working "
        f"directory {root_resolved}; refusing to read/write outside the project"
    )


# --------------------------------------------------------------------------- #
# Markdown → content model (the map step)
# --------------------------------------------------------------------------- #
def parse_markdown(text: str) -> dict:
    """Project a Markdown artifact onto the xlsx content model.

    Mapping: front-matter ``key: value`` lines → ``scalars`` (each
    written into a single-cell *named range* of the same name); the first
    Markdown table → ``table`` (``{header, rows}``, written into an Excel
    *Table* data region). Pure transform — no I/O.
    """
    scalars: dict[str, str] = {}
    table: dict | None = None
    lines = text.splitlines()
    i = 0

    if lines and lines[0].strip() == "---":
        j = 1
        while j < len(lines) and lines[j].strip() != "---":
            raw = lines[j]
            if ":" in raw:
                key, _, value = raw.partition(":")
                scalars[key.strip()] = value.strip().strip("\"'")
            j += 1
        i = j + 1 if j < len(lines) else len(lines)

    def cell_split(row: str) -> list[str]:
        return [c.strip() for c in row.strip().strip("|").split("|")]

    while i < len(lines):
        stripped = lines[i].strip()
        if (
            table is None
            and stripped.startswith("|")
            and i + 1 < len(lines)
            and "-" in lines[i + 1]
            and set(lines[i + 1].strip().replace("|", "").replace(":", "").strip()) <= {"-", " "}
            and lines[i + 1].strip().replace("|", "").strip()
        ):
            header = cell_split(stripped)
            rows = []
            i += 2
            while i < len(lines) and lines[i].strip().startswith("|"):
                rows.append(cell_split(lines[i]))
                i += 1
            table = {"header": header, "rows": rows}
            continue
        i += 1

    return {"scalars": scalars, "table": table}


# --------------------------------------------------------------------------- #
# openpyxl helpers (imported lazily so --check runs without the library)
# --------------------------------------------------------------------------- #
def _strip_abs(coord: str) -> str:
    return coord.replace("$", "")


def inspect_template(template: Path) -> list[dict]:
    """Return the workbook's fill-points: defined names + Excel Tables."""
    from openpyxl import load_workbook

    wb = load_workbook(str(template))
    fillpoints: list[dict] = []
    for name in wb.defined_names:
        dn = wb.defined_names[name]
        ref = ";".join(f"{title}!{coord}" for title, coord in dn.destinations)
        fillpoints.append({"kind": "defined_name", "name": name, "ref": ref})
    for ws in wb.worksheets:
        # TableList.values() yields Table objects; .items() yields raw (name, ref).
        for tbl in ws.tables.values():
            fillpoints.append({"kind": "table", "name": tbl.name, "ref": f"{ws.title}!{tbl.ref}"})
    return fillpoints


def _write_named_scalars(wb, scalars: dict) -> tuple[int, list[str]]:
    """Write each front-matter scalar into the single-cell named range it matches."""
    filled = 0
    warnings: list[str] = []
    for key, value in scalars.items():
        if key not in wb.defined_names:
            continue
        dests = list(wb.defined_names[key].destinations)
        # A single cell yields one destination whose coord has no ":" range
        # separator; a multi-cell name yields a range coord (or several dests).
        if len(dests) != 1 or ":" in dests[0][1]:
            warnings.append(f"named range '{key}' is not a single cell; skipped")
            continue
        title, coord = dests[0]
        wb[title][_strip_abs(coord)] = value
        filled += 1
    return filled, warnings


def _write_table(wb, table: dict) -> tuple[int, list[str]]:
    """Write the Markdown table into the first Excel Table's data region.

    Columns align by header text where they match, else by position. Rows fill
    the table's *existing* data region; the range is never resized (resizing
    could break a chart that reads it). Overflow rows are truncated with a
    warning.
    """
    from openpyxl.utils.cell import range_boundaries, get_column_letter

    target_ws = target_tbl = None
    for ws in wb.worksheets:
        if ws.tables:
            target_ws = ws
            target_tbl = next(iter(ws.tables.values()))
            break
    if target_tbl is None:
        return 0, []

    min_col, min_row, max_col, max_row = range_boundaries(target_tbl.ref)
    tbl_header = [target_ws.cell(row=min_row, column=c).value for c in range(min_col, max_col + 1)]

    # Column index in the template table for each Markdown column (by header text).
    md_header = table["header"]
    col_for_md = {}
    for mi, mh in enumerate(md_header):
        if mh in tbl_header:
            col_for_md[mi] = min_col + tbl_header.index(mh)
        elif mi < (max_col - min_col + 1):
            col_for_md[mi] = min_col + mi  # positional fallback

    data_rows = max_row - min_row  # rows under the header
    warnings: list[str] = []
    if len(table["rows"]) > data_rows:
        warnings.append(
            f"Markdown table has {len(table['rows'])} rows but the Excel Table "
            f"'{target_tbl.name}' has room for {data_rows}; extra rows were truncated "
            "(the table range is not resized, to avoid breaking charts)"
        )

    filled = 0
    for r, row in enumerate(table["rows"][:data_rows]):
        for mi, value in enumerate(row):
            col = col_for_md.get(mi)
            if col is not None:
                target_ws.cell(row=min_row + 1 + r, column=col, value=value)
        filled += 1
    if filled:
        # touch get_column_letter so the import is exercised on the happy path
        _ = get_column_letter(min_col)
    return filled, warnings


def render_xlsx(text: str, template: Path, output: Path) -> RenderResult:
    """Fill ``template`` from the Markdown in ``text``; write ``output``.

    Writes only into named ranges and Excel-Table data cells; never touches
    chart/shape objects. If the workbook has no fill-points, returns guidance
    instead of writing (never a silent convert).
    """
    from openpyxl import load_workbook

    model = parse_markdown(text)
    wb = load_workbook(str(template))

    has_names = len(list(wb.defined_names)) > 0
    has_tables = any(ws.tables for ws in wb.worksheets)
    if not has_names and not has_tables:
        return RenderResult(written=False, filled=0, warnings=[], guidance=GUIDANCE_NO_FILLPOINTS)

    filled_s, warn_s = _write_named_scalars(wb, model["scalars"])
    filled_t, warn_t = (0, [])
    if model["table"] is not None:
        filled_t, warn_t = _write_table(wb, model["table"])

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output))
    return RenderResult(
        written=True, filled=filled_s + filled_t, warnings=warn_s + warn_t, guidance=None
    )


def render_default_xlsx(text: str, output: Path) -> RenderResult:
    """Opt-out path: write the Markdown into a bare openpyxl workbook.

    Used only when the user explicitly declines a branded template. The result
    is unbranded — front-matter scalars go into a key/value block, the first
    Markdown table follows below. Not the primary path: branding requires a
    template with named ranges / Tables (``render_xlsx``).
    """
    from openpyxl import Workbook

    model = parse_markdown(text)
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    row = 1
    for key, value in model["scalars"].items():
        ws.cell(row=row, column=1, value=key)
        ws.cell(row=row, column=2, value=value)
        row += 1
    if model["scalars"] and model["table"]:
        row += 1  # blank spacer between the scalar block and the table

    table = model["table"]
    if table:
        for c, h in enumerate(table["header"], start=1):
            ws.cell(row=row, column=c, value=h)
        for r, data_row in enumerate(table["rows"], start=1):
            for c, cell in enumerate(data_row, start=1):
                ws.cell(row=row + r, column=c, value=cell)

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output))
    return RenderResult(
        written=True,
        filled=0,
        warnings=["no template supplied — produced an UNBRANDED .xlsx via openpyxl"],
        guidance=None,
    )


# --------------------------------------------------------------------------- #
# CLI
# --------------------------------------------------------------------------- #
def cmd_check() -> int:
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        print(f"openpyxl is not installed. Run:\n    {PIP_INSTALL}", file=sys.stderr)
        return 2
    return 0


def cmd_inspect(args) -> int:
    root = Path.cwd()
    template = confine(Path(args.template), root)
    if not template.is_file():
        print(f"ERROR: template not found: {args.template}", file=sys.stderr)
        return 1
    fillpoints = inspect_template(template)
    if not fillpoints:
        print(f"GUIDANCE: {GUIDANCE_NO_FILLPOINTS}")
        return 0
    for fp in fillpoints:
        print(f"FILLPOINTS: kind={fp['kind']} name={fp['name']} ref={fp['ref']}")
    return 0


def cmd_render(args) -> int:
    root = Path.cwd()
    md_path = confine(Path(args.markdown), root)
    if not md_path.is_file():
        print(f"ERROR: markdown not found: {args.markdown}", file=sys.stderr)
        return 1

    output = Path(args.output) if args.output else Path(md_path.stem + ".xlsx")
    output = confine(output, root)
    text = md_path.read_text(encoding="utf-8")

    if not args.template:
        # Explicit opt-out (the agent omits --template only after the user
        # declines a branded template, per SKILL.md): bare openpyxl workbook.
        result = render_default_xlsx(text, output)
    else:
        template = confine(Path(args.template), root)
        if not template.is_file():
            print(f"ERROR: template not found: {args.template}", file=sys.stderr)
            return 1
        result = render_xlsx(text, template, output)
    if result.guidance:
        print(f"GUIDANCE: {result.guidance}")
        return 0
    for w in result.warnings:
        print(f"WARNING: {w}")
    print(f"OUTPUT: {output}")
    print(f"FILLED: {result.filled}")
    return 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Fill an Excel template from Markdown.")
    parser.add_argument("--check", action="store_true", help="probe openpyxl; exit 0/2")
    sub = parser.add_subparsers(dest="verb")

    p_inspect = sub.add_parser("inspect", help="print named ranges + Excel Tables")
    p_inspect.add_argument("template")

    p_render = sub.add_parser("render", help="fill a template and write a .xlsx")
    p_render.add_argument("markdown")
    p_render.add_argument("--template", help="workbook with named ranges/tables to fill")
    p_render.add_argument("--output", help="output path (default: <basename>.xlsx in CWD)")
    return parser


def main(argv: list[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)

    if args.check:
        return cmd_check()

    rc = cmd_check()
    if rc != 0:
        return rc

    try:
        if args.verb == "inspect":
            return cmd_inspect(args)
        if args.verb == "render":
            return cmd_render(args)
    except ValueError as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1

    parser.print_help()
    return 1


if __name__ == "__main__":
    raise SystemExit(main())
