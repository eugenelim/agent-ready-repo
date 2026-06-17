"""TDD suite for the markdown-to-xlsx renderer.

Covers the contracts the spec names: fill-point enumeration (named ranges +
Excel Tables), Markdown→content-model mapping, the no-fill-points guidance, the
data-ranges-only path leaving a chart intact, the Tier-1 `--check` probe, and
output-path confinement — plus an end-to-end render that re-opens the produced
.xlsx and asserts the filled values are present.

The fixture workbook is built with openpyxl at test time rather than committed
as an opaque binary; the builder below is its authorship. Run with
`python -m pytest` from this directory.
"""

from __future__ import annotations

import subprocess
import sys
from pathlib import Path

import pytest

import render

HERE = Path(__file__).resolve().parent

MARKDOWN = """\
---
report_title: Q3 Review
prepared_by: Finance
---

# Numbers

| Metric | Value |
| --- | --- |
| ARR | 4.2M |
| Churn | 1.1% |
"""


def _build_template(path: Path, with_chart: bool = False) -> Path:
    """Author a .xlsx template with a named range, an Excel Table, optional chart."""
    from openpyxl import Workbook
    from openpyxl.workbook.defined_name import DefinedName
    from openpyxl.worksheet.table import Table

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    # Title cell + a single-cell named range pointing at it.
    ws["D1"] = ""
    wb.defined_names.add(DefinedName("report_title", attr_text="Data!$D$1"))
    # A data block headed by Metric/Value with two empty data rows, as a Table.
    ws["A1"], ws["B1"] = "Metric", "Value"
    for r in (2, 3):
        ws.cell(row=r, column=1, value="")
        ws.cell(row=r, column=2, value="")
    ws.add_table(Table(displayName="metrics", ref="A1:B3"))

    if with_chart:
        from openpyxl.chart import BarChart, Reference

        chart = BarChart()
        chart.add_data(Reference(ws, min_col=2, min_row=1, max_row=3), titles_from_data=True)
        ws.add_chart(chart, "F2")

    wb.save(str(path))
    return path


def _build_empty_template(path: Path) -> Path:
    from openpyxl import Workbook

    wb = Workbook()
    wb.active["A1"] = "no fill-points here"
    wb.save(str(path))
    return path


# --------------------------------------------------------------------------- #
# inspect — named ranges + Excel Tables
# --------------------------------------------------------------------------- #
def test_inspect_returns_defined_names_and_tables(tmp_path):
    template = _build_template(tmp_path / "template.xlsx")
    fillpoints = render.inspect_template(template)
    kinds = {(fp["kind"], fp["name"]) for fp in fillpoints}
    assert ("defined_name", "report_title") in kinds
    assert ("table", "metrics") in kinds


# --------------------------------------------------------------------------- #
# map — Markdown → content model
# --------------------------------------------------------------------------- #
def test_map_projects_frontmatter_and_table():
    model = render.parse_markdown(MARKDOWN)
    # front-matter → single-cell named-range scalars
    assert model["scalars"]["report_title"] == "Q3 Review"
    assert model["scalars"]["prepared_by"] == "Finance"
    # a Markdown table → a Table data region
    assert model["table"]["header"] == ["Metric", "Value"]
    assert model["table"]["rows"] == [["ARR", "4.2M"], ["Churn", "1.1%"]]


# --------------------------------------------------------------------------- #
# no-fill-points guidance
# --------------------------------------------------------------------------- #
def test_empty_markdown_parses_without_crash():
    assert render.parse_markdown("") == {"scalars": {}, "table": None}


def test_multi_cell_named_range_is_skipped_with_warning(tmp_path):
    from openpyxl import Workbook
    from openpyxl.workbook.defined_name import DefinedName

    wb = Workbook()
    wb.active.title = "Data"
    wb.defined_names.add(DefinedName("report_title", attr_text="Data!$A$1:$A$2"))
    wb.save(str(tmp_path / "t.xlsx"))

    out = tmp_path / "out.xlsx"
    result = render.render_xlsx("---\nreport_title: Q3\n---\n", tmp_path / "t.xlsx", out)
    assert result.written is True
    assert any("not a single cell" in w for w in result.warnings)


def test_overflow_rows_truncated_with_warning(tmp_path):
    from openpyxl import load_workbook

    template = _build_template(tmp_path / "template.xlsx")  # Table A1:B3 → 2 data rows
    md = (
        "---\ntitle: x\n---\n\n# N\n\n| Metric | Value |\n| --- | --- |\n"
        "| ARR | 1 |\n| Churn | 2 |\n| NRR | 3 |\n"
    )
    out = tmp_path / "out.xlsx"
    result = render.render_xlsx(md, template, out)
    assert any("truncated" in w for w in result.warnings)
    ws = load_workbook(str(out))["Data"]
    assert ws["A2"].value == "ARR" and ws["A3"].value == "Churn"
    assert ws["A4"].value in (None, "")  # the 3rd row did not overflow the range


def test_render_default_writes_unbranded_xlsx(tmp_path):
    from openpyxl import load_workbook

    out = tmp_path / "out.xlsx"
    result = render.render_default_xlsx(MARKDOWN, out)
    assert result.written is True
    assert any("UNBRANDED" in w for w in result.warnings)
    ws = load_workbook(str(out))["Data"]
    cells = [c.value for row in ws.iter_rows() for c in row]
    assert "Q3 Review" in cells and "ARR" in cells and "4.2M" in cells


def test_no_fillpoints_yields_guidance(tmp_path):
    template = _build_empty_template(tmp_path / "empty.xlsx")
    assert render.inspect_template(template) == []
    out = tmp_path / "out.xlsx"
    result = render.render_xlsx(MARKDOWN, template, out)
    assert result.written is False
    assert result.guidance and "named range" in result.guidance
    assert not out.exists()


# --------------------------------------------------------------------------- #
# data-ranges-only: a chart survives the render
# --------------------------------------------------------------------------- #
def test_render_preserves_chart_and_writes_only_data(tmp_path):
    from openpyxl import load_workbook

    template = _build_template(tmp_path / "template.xlsx", with_chart=True)
    out = tmp_path / "out.xlsx"
    result = render.render_xlsx(MARKDOWN, template, out)
    assert result.written is True

    wb = load_workbook(str(out))
    ws = wb["Data"]
    # the chart object is still present (the path never touched it)
    assert len(ws._charts) == 1
    # the data ranges were written
    assert ws["D1"].value == "Q3 Review"
    assert ws["A2"].value == "ARR" and ws["B2"].value == "4.2M"


# --------------------------------------------------------------------------- #
# --check — Tier-1 probe
# --------------------------------------------------------------------------- #
def test_check_exit_zero_when_present():
    assert render.cmd_check() == 0


def test_check_exit_two_when_absent(monkeypatch):
    monkeypatch.setitem(sys.modules, "openpyxl", None)
    assert render.cmd_check() == 2


# --------------------------------------------------------------------------- #
# confinement
# --------------------------------------------------------------------------- #
def test_confine_rejects_parent_traversal(tmp_path):
    root = tmp_path / "workdir"
    root.mkdir()
    with pytest.raises(ValueError):
        render.confine(root / ".." / "evil.xlsx", root)


def test_confine_rejects_symlink_escape(tmp_path):
    root = tmp_path / "workdir"
    root.mkdir()
    outside = tmp_path / "outside"
    outside.mkdir()
    link = root / "link"
    link.symlink_to(outside, target_is_directory=True)
    with pytest.raises(ValueError):
        render.confine(link / "evil.xlsx", root)


def test_confine_rejects_sibling_prefix(tmp_path):
    root = tmp_path / "workdir"
    root.mkdir()
    (tmp_path / "workdir-evil").mkdir()
    with pytest.raises(ValueError):
        render.confine(tmp_path / "workdir-evil" / "out.xlsx", root)


# --------------------------------------------------------------------------- #
# end-to-end render (manual-QA mode, exercised as an integration test)
# --------------------------------------------------------------------------- #
def test_render_fills_template_and_reopened_values_present(tmp_path):
    from openpyxl import load_workbook

    _build_template(tmp_path / "template.xlsx")
    (tmp_path / "report.md").write_text(MARKDOWN, encoding="utf-8")

    result = subprocess.run(
        [sys.executable, str(HERE / "render.py"), "render", "report.md",
         "--template", "template.xlsx", "--output", "out.xlsx"],
        cwd=tmp_path, capture_output=True, text=True,
    )
    assert result.returncode == 0, result.stderr
    assert "OUTPUT:" in result.stdout
    out = tmp_path / "out.xlsx"
    assert out.is_file()

    ws = load_workbook(str(out))["Data"]
    assert ws["D1"].value == "Q3 Review"
    assert ws["A2"].value == "ARR" and ws["B2"].value == "4.2M"
    assert ws["A3"].value == "Churn" and ws["B3"].value == "1.1%"
